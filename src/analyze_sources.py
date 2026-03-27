#!/usr/bin/env python3
"""
Analyze source files for Quotation App Project
Generates insights about structure and data dependencies.
"""

import sys
import os
from pathlib import Path

# Add parent directory to path so we can import our modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: Missing required packages")
    print("Install with: pip install pandas openpyxl")
    sys.exit(1)


def analyze_jsx_structure():
    """Extract key info from JSX calculator file"""
    print("=" * 70)
    print("JSX CALCULATOR ANALYSIS")
    print("=" * 70)

    jsx_path = Path(__file__).parent / "kalkulator_wycen.jsx"

    if not jsx_path.exists():
        print(f"ERROR: JSX file not found at {jsx_path}")
        return

    with open(jsx_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Count operation definitions
    op_defs_start = content.find("const OP_DEFS = [")
    op_defs_end = content.find("];", op_defs_start)
    op_defs_section = content[op_defs_start:op_defs_end]

    operation_count = op_defs_section.count('id: "')
    machine_count = op_defs_section.count('k: "')

    print(f"✓ Operations defined: {operation_count}")
    print(f"✓ Machines referenced: {machine_count}")
    print(f"✓ File size: {len(content)} characters")

    # Extract machine rates
    rates_start = content.find("const RATES = {")
    rates_end = content.find("};", rates_start)
    rates_section = content[rates_start:rates_end + 2]

    print("\nMachine rates found:")
    print(rates_section[:300] + "..." if len(rates_section) > 300 else rates_section)


def analyze_excel_files():
    """Analyze both Excel files"""
    project_root = Path(__file__).parent.parent
    data_dir = project_root / "data"

    if not data_dir.exists():
        print(f"ERROR: Data directory not found at {data_dir}")
        return

    print("\n" + "=" * 70)
    print("EXCEL FILES ANALYSIS")
    print("=" * 70)

    for excel_file in data_dir.glob("*.xlsx"):
        print(f"\nFile: {excel_file.name}")
        print("-" * 70)

        try:
            wb = load_workbook(excel_file, read_only=True, data_only=True)
            print(f"✓ Sheets: {len(wb.sheetnames)} - {wb.sheetnames}")

            # Get basic info from first sheet
            first_sheet = wb.sheetnames[0]
            df = pd.read_excel(excel_file, sheet_name=first_sheet, header=None)
            print(f"✓ First sheet dimensions: {df.shape[0]} rows x {df.shape[1]} columns")

            # Preview first 5 rows
            print(f"\nFirst 5 rows of '{first_sheet}':")
            for i in range(min(5, len(df))):
                row = df.iloc[i].dropna().head(5)
                if not row.empty:
                    print(f"  Row {i}: {row.tolist()[:3]}...")

        except Exception as e:
            print(f"✗ Error: {str(e)}")


def generate_project_summary():
    """Generate summary for planning"""
    print("\n" + "=" * 70)
    print("PROJECT IMPLEMENTATION SUMMARY")
    print("=" * 70)

    print("""
RECOMMENDED PYTHON PACKAGE STRUCTURE:
├── src/
│   └── quotation_app/
│       ├── __init__.py
│       ├── calculator.py      # Main calculation engine
│       ├── operations.py        # Operation definitions & formulas
│       ├── materials.py         # Material database
│       ├── machines.py          # Machine rates
│       └── utils.py            # Helper functions
├── data/
│   ├── kalkulator_wycen.jsx
│   └── *.xlsx                  # Source data
├── tests/
│   └── test_operations.py      # Validation tests
├── venv/                       # Virtual environment
├── requirements.txt
├── README.md
└── setup.py                    # Package installer

KEY COMPONENTS TO IMPLEMENT:
1. Operations module - Translate 15+ JSX operations to Python classes
2. Calculation engine - Replicate complex formulas from JSX
3. Material parser - Extract data from Excel price list
4. Validation layer - Compare Python vs Excel outputs
5. CLI interface - User-friendly command-line tool
6. Tests - Automated comparison with source files

NEXT STEPS (1-hour chunks):
→ Design calculator.py interface
→ Implement first operation (Druk Sitowy)
→ Test against Excel data
→ Add second operation
→ Build material parser
→ Create CLI for first feature
→ Add validation layer
""")


if __name__ == "__main__":
    print("Quotation App - Source Files Analysis")
    print("=" * 70)

    analyze_jsx_structure()
    analyze_excel_files()
    generate_project_summary()

    print("\n" + "=" * 70)
    print("Analysis complete!")
    print("Run this script anytime to refresh understanding of source files.")
    print("=" * 70)
